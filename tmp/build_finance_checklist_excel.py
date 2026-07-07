from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from pathlib import Path


rows = [
    ("Group 1", "1.1", "M1", "Schema migration: add 5 columns to refunds table", "", "Done", "Refund workflow columns added and aligned with finance flow."),
    ("Group 1", "1.2", "M1", "Create platform_settings table + seed refund threshold (200,000 paise)", "", "Done", "Platform settings table and refund threshold seed are implemented."),
    ("Group 1", "1.3", "M10", "Create financial_audit_log table + REVOKE DELETE/UPDATE at DB level", "", "Done", "Implemented on shared audit_logs with append-only DB trigger and finance semantics."),
    ("Group 1", "1.4", "M7", "Create legal_contacts table", "", "Done", "Legal contacts schema and history model are present."),
    ("Group 1", "1.5", "M9", "Create module_access table", "", "Done", "Module access table exists and is used for finance gating."),
    ("Group 1", "1.6", "M9", "Create pl_manual_entries table", "", "Done", "Manual P&L entries table exists with lock-compatible fields."),
    ("Group 1", "1.7", "M5", "Create brand_tds_tracking table", "", "Done", "TDS tracking table and FY rollover flow are implemented."),
    ("Group 1", "1.8", "M6", "Create user_consents table", "", "Done", "Consent storage schema is implemented."),
    ("Group 1", "1.9", "M6", "Create data_deletion_requests table", "", "Done", "Deletion request schema and workflow fields are implemented."),
    ("Group 1", "1.10", "M3", "Create commission_rules table + brand_payout_config table", "", "Done", "Commission rules and payout config foundations are implemented."),
    ("Group 1", "1.11", "M2", "Create carrier_fee_schedule + cod_reconciliation + cod_reconciliation_runs tables", "", "Done", "COD reconciliation schema set is implemented."),
    ("Group 1", "1.12", "M4", "Create hsn_master table", "", "Done", "HSN master schema exists and is wired into GST flow."),
    ("Group 2", "2.1", "M1", "Admin UI: reason_code dropdown wired to reason_master + cost_allocation auto-suggest", "1.1, 1.3", "Done", "Refund UI supports reason selection and cost allocation suggestions."),
    ("Group 2", "2.2", "M1", "Approval gate logic: block processing above threshold, alert approver", "2.1", "Done", "Threshold approval gate and awaiting-approval flow are implemented."),
    ("Group 2", "2.3", "M1", "Razorpay refund hook: auto-write audit log on every state change", "1.3", "Done", "Refund transitions and webhook outcomes now write finance audit entries."),
    ("Group 2", "2.4", "M1", "Shiprocket reverse pickup auto-trigger for brand_fault / renivet_fault", "2.2", "Done", "Reverse pickup creation is wired for the required fault categories."),
    ("Group 2", "2.5", "M10", "Wire audit hooks: payout events, commission changes, bank detail changes", "1.3", "Done", "Finance audit hooks cover payout, commission, and bank-detail events."),
    ("Group 2", "2.6", "M10", "Wire audit hooks: GST exports, P&L edits, module access, data deletion", "1.3", "Done", "Audit hooks cover GST export, P&L edits, module access, and deletion flows."),
    ("Group 2", "2.7", "M7", "Admin UI: GRO edit form in Settings -> Legal Contacts (with audit log + alert)", "1.4, 2.5", "Done", "GRO admin update flow is implemented with audit and alert behavior."),
    ("Group 2", "2.8", "M7", "Website: Contact Us GRO section (dynamic from DB)", "1.4", "Done", "Public GRO content is served from legal contact records."),
    ("Group 2", "2.9", "M7", "Website: Footer Grievance Redressal link + ToS clause added", "2.8", "Done", "Footer and ToS grievance additions were implemented."),
    ("Group 2", "2.10", "M6", "Clerk signup: consent modal with required + optional consents", "1.8", "Partial", "Consent storage is implemented, but end-to-end signup modal validation still needs final QA."),
    ("Group 2", "2.11", "M6", "Account Settings: consent management page + Brevo/AiSensy propagation on opt-out", "2.10", "Partial", "Consent management flow exists, but integration completeness should be verified in live testing."),
    ("Group 3", "3.1", "", "Delhivery API: weekly COD fee schedule fetch + upsert to carrier_fee_schedule", "1.11", "Done", "COD fee sync route and storage are implemented."),
    ("Group 3", "3.2", "", "Delhivery API: daily remittance sync cron + auto-categorization logic", "3.1", "Done", "Daily remittance sync and discrepancy categorization are implemented."),
    ("Group 3", "3.3", "", "Shiprocket remittance sync", "1.11", "Partial", "Primary COD implementation is live; Shiprocket remittance support should be treated as follow-up verification."),
    ("Group 3", "3.4", "", "Admin dashboard: COD reconciliation table with status filters + one-click actions", "3.2, 3.3", "Done", "COD dashboard workspace, filters, and actions are implemented."),
    ("Group 3", "3.5", "", "Overdue/critical alert crons to AJ", "3.2", "Done", "Alerting for COD failures and overdue states is wired."),
    ("Group 3", "3.6", "", "Write-off flow: mandatory reason + proof upload + audit log", "3.4, 2.5", "Done", "COD write-off requires proof and logs to finance audit."),
    ("Group 4", "4.1", "", "Commission rule resolution engine (brand + category lookup, priority, fallback)", "1.10", "Done", "Commission resolution engine is implemented."),
    ("Group 4", "4.2", "", "Payout calculation engine (full formula per brand per cycle)", "4.1, 2.3", "Done", "Cycle calculation supports commission, returns, holdback, overrides, and TDS."),
    ("Group 4", "4.3", "", "Admin payout dashboard: per-brand review cards with expandable line items", "4.2", "Done", "Payout review workspace is implemented."),
    ("Group 4", "4.4", "", "Override flow: form + proof upload + dual-approval for overrides >Rs500", "4.3, 2.5", "Done", "Override flow supports proof and separate approver logic."),
    ("Group 4", "4.5", "", "Payment execution: Razorpay Payout API for razorpay_route brands", "4.3", "Done", "Automated payout execution path is implemented."),
    ("Group 4", "4.6", "", "Payment execution: Manual NEFT txn ID capture flow for manual_neft brands", "4.3", "Done", "Manual NEFT execution and confirmation flow is implemented."),
    ("Group 4", "4.7", "", "Brand statement PDF generation (all sections including overrides, TDS)", "4.2", "Done", "Statement PDF route exists and includes payout summary details."),
    ("Group 4", "4.8", "", "Payout notifications: cycle day alert to AJ + execution confirmation to brand", "4.5, 4.6", "Done", "Cycle alerts and payout event notifications are wired."),
    ("Group 4", "4.9", "", "TDS deduction integrated into payout calculation + brand statement display", "4.2, 1.7", "Done", "TDS is calculated during payout and surfaced in payout data."),
    ("Group 4", "4.10", "", "Quarterly TDS summary export CSV for RegisterKaro", "4.9", "Done", "Quarterly TDS export endpoint is implemented."),
    ("Group 5", "5.0", "", "Prerequisite action (not IT): Confirm CSV format with RegisterKaro before any build starts", "-", "External", "Business-side dependency; not a code deliverable."),
    ("Group 5", "5.1", "", "Admin HSN import tool: CSV upload to populate hsn_master", "1.12", "Done", "HSN admin import/manage flow is implemented."),
    ("Group 5", "5.2", "", "Products table: add hsn_master FK, link existing HSN codes", "5.1", "Partial", "HSN integration is implemented at report level; existing product linkage should be verified against production data."),
    ("Group 5", "5.3", "", "GST rate auto-lookup + IGST/CGST/SGST split logic based on customer state", "5.2", "Done", "GST split logic is implemented in report generation."),
    ("Group 5", "5.4", "", "Pre-flight validation: flag missing HSN codes and brand GSTINs", "5.3", "Done", "GST preview blocks export when required data is missing."),
    ("Group 5", "5.5", "", "Combined GST + TCS CSV generator (two sections, one file)", "5.3, 5.4", "Done", "Combined GST/TCS export generation is implemented."),
    ("Group 5", "5.6", "", "Admin: Finance -> GST Report UI with preview summary and download", "5.5", "Done", "GST workspace includes preview, validation, and export flows."),
    ("Group 6", "6.1", "", "Account Settings: 'Delete My Account' flow with email verification", "1.9", "Partial", "Deletion request and verification model exist; final account-settings UX should be validated end to end."),
    ("Group 6", "6.2", "", "Admin: data deletion request review + execute panel", "1.9, 2.6", "Done", "Admin review and execution flow is implemented."),
    ("Group 6", "6.3", "", "Anonymization script: all 10 steps in atomic transaction", "6.2", "Done", "Deletion executor and anonymization flow are implemented."),
    ("Group 6", "6.4", "", "SLA cron: day 5 alert + day 7 escalation to AJ for pending requests", "1.9", "Done", "Deletion SLA sweep cron is implemented."),
    ("Group 7", "7.1", "", "Module access middleware check on monthly_pl API routes", "1.5", "Done", "Monthly P&L APIs enforce explicit module access."),
    ("Group 7", "7.2", "", "Admin Settings: Module Access management UI (super-admin only)", "1.5, 2.6", "Done", "AJ-only grant/revoke UI is implemented."),
    ("Group 7", "7.3", "", "Razorpay API: monthly gateway fee fetch", "-", "Partial", "P&L currently derives gateway fees; direct monthly settlement import still needs hard validation."),
    ("Group 7", "7.4", "", "Delhivery + Shiprocket API: monthly shipping invoice totals", "3.1, 3.3", "Partial", "P&L shipping cost layer is present, but invoice-grade totals should be verified against carrier exports."),
    ("Group 7", "7.5", "", "P&L data aggregation service (combines all auto sources)", "7.3, 7.4", "Partial", "Aggregation service exists, but some auto-sources are still derived rather than provider-perfect."),
    ("Group 7", "7.6", "", "P&L dashboard UI: auto fields with source badges + manual entry fields", "7.5, 1.6", "Done", "Monthly P&L dashboard UI is implemented."),
    ("Group 7", "7.7", "", "Lock month mechanism + pl_snapshots for historical preservation", "7.6, 2.6", "Done", "Lock/unlock and snapshot preservation are implemented."),
    ("Group 7", "7.8", "", "PDF export of locked P&L month", "7.7", "Done", "Locked month PDF export route is implemented."),
    ("Group 8", "8.1", "", "Audit log viewer UI with filters, diff viewer, CSV export", "All hook tasks (2.5, 2.6, and later groups)", "Done", "Finance audit log viewer, filters, JSON diff view, and CSV export are implemented."),
    ("Group 8", "8.2", "", "Attachment viewer: display proof docs from Uploadthing inline", "8.1", "Done", "Audit viewer exposes attachment links for proof files."),
]


status_fill = {
    "Done": "DCFCE7",
    "Partial": "FEF3C7",
    "External": "E5E7EB",
    "Not Done": "FEE2E2",
}


def build_workbook(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist Status"
    summary = wb.create_sheet("Summary")

    title_fill = PatternFill("solid", fgColor="0F172A")
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Finance Build Checklist Status"
    ws["A2"] = "Status based on current implementation in repo and recent completion sweep."
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A2"].font = Font(color="334155", italic=True, size=10)
    ws["A1"].fill = title_fill
    ws["A2"].fill = PatternFill("solid", fgColor="EFF6FF")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Group", "Task ID", "Module", "Task", "Needs", "Status", "Short Description"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="1E3A8A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_group = None
    row_idx = 5
    for row in rows:
        group, task_id, module, task, needs, status, note = row
        if group != current_group:
            current_group = group
        values = [group, task_id, module, task, needs, status, note]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 6:
                cell.fill = PatternFill("solid", fgColor=status_fill[status])
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

    widths = {
        "A": 14,
        "B": 10,
        "C": 10,
        "D": 58,
        "E": 24,
        "F": 12,
        "G": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{row_idx-1}"

    done_count = sum(1 for r in rows if r[5] == "Done")
    partial_count = sum(1 for r in rows if r[5] == "Partial")
    external_count = sum(1 for r in rows if r[5] == "External")
    total_count = len(rows)

    summary["A1"] = "Completion Summary"
    summary.merge_cells("A1:D1")
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].fill = title_fill
    summary["A1"].alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Total tasks", total_count),
        ("Done", done_count),
        ("Partial", partial_count),
        ("External", external_count),
        ("Completion % (Done only)", done_count / total_count),
    ]

    summary["A3"] = "Metric"
    summary["B3"] = "Value"
    summary["A3"].fill = header_fill
    summary["B3"].fill = header_fill
    summary["A3"].font = Font(bold=True, color="1E3A8A")
    summary["B3"].font = Font(bold=True, color="1E3A8A")
    summary["A3"].border = border
    summary["B3"].border = border

    for idx, (label, value) in enumerate(summary_rows, start=4):
        summary[f"A{idx}"] = label
        summary[f"B{idx}"] = value
        summary[f"A{idx}"].border = border
        summary[f"B{idx}"].border = border
        if isinstance(value, float):
            summary[f"B{idx}"].number_format = "0.0%"

    summary["A10"] = "Notes"
    summary["A10"].fill = header_fill
    summary["A10"].font = Font(bold=True, color="1E3A8A")
    summary["A10"].border = border
    notes = [
        "Done = implemented in current repo state based on recent completion sweep.",
        "Partial = code exists but still needs end-to-end verification or external-system hardening.",
        "External = not an IT-owned deliverable.",
        "Finance audit log is implemented via the shared audit_logs table, not a separate financial_audit_log table.",
    ]
    for idx, note in enumerate(notes, start=11):
        summary[f"A{idx}"] = note
        summary[f"A{idx}"].border = border

    summary.column_dimensions["A"].width = 80
    summary.column_dimensions["B"].width = 18

    wb.save(output_path)


def verify_workbook(output_path: Path) -> None:
    wb = load_workbook(output_path)
    assert "Checklist Status" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    ws = wb["Checklist Status"]
    assert ws["A1"].value == "Finance Build Checklist Status"
    assert ws.max_row >= len(rows) + 4


if __name__ == "__main__":
    output_dir = Path(r"C:\Personal Projects\renivet-marketplace\outputs\finance-checklist")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "finance_build_checklist_status.xlsx"
    build_workbook(output_path)
    verify_workbook(output_path)
    print(output_path)
